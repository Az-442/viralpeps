#!/usr/bin/env python3
"""Build the finalized TrustScore outreach tracker .xlsx with emails + Status dropdown."""
import json, subprocess
from pathlib import Path

d = json.load(open('/Users/time4you/viralpeps/src/data/vendors.json'))
vendors = d if isinstance(d, list) else d.get('vendors', d)
emails = json.load(open('/tmp/supplier_emails.json'))

OFFLINE = {'Sterling Peptides','Brit Peptides','Chroma Peptides','Chilton Labs','Bio Pulse Peptides'}
# contact-only notes for missing
NOTES = {
    'Express Peptides': 'WhatsApp 07405 995349',
    'XL Peptides': 'WhatsApp only',
    'Peptides UK': 'WhatsApp only',
    'Raw Peptides': 'WhatsApp only',
    'PeptX': 'form only',
    'CMSR Labs': 'form only',
    'Astra Labs': 'form only',
    'PGNA Labs': 'form only',
    'Bio Peptides UK': 'form only',
    'RETA UK': 'form only',
    'United Peptides': 'OWN BUSINESS',
}

def entity_points(et, pm):
    if pm in ('card','bank'): return 25
    if et=='ltd': return 20
    if et=='sole_trader': return 10
    return 0

def score(x):
    a = x.get('_autoChecks') or {}; s = 0
    if a.get('coa') is True: s += 25
    if a.get('ruo') is True: s += 5
    if a.get('reviews') is True: s += 10
    if a.get('shipping') is True: s += 5
    if a.get('contact') is True: s += 10
    if x.get('embedded') is True or x.get('domainVerified') is True: s += 20
    ep = entity_points(x.get('entityType'), x.get('paymentMethod'))
    if ep > 0: s += ep
    return min(s, 100)

rows = []
for x in vendors:
    nm = x.get('name','')
    live = (x.get('_autoChecks') or {}).get('live')
    seg = 'Consultancy/compliance (offline)' if nm in OFFLINE else ('Tier A - best fit' if score(x) >= 55 else 'Tier B - good fit')
    email = emails.get(nm, '')
    note = NOTES.get(nm, '')
    if not email and note: note = note  # already set
    elif not email and not note: note = 'no email found'
    rows.append([nm, x.get('website',''), score(x), seg, email, note, bool(live)])

# sort: live first, offline last; score desc
rows.sort(key=lambda r: (0 if r[6] else 1, -r[2]))

ox_script = r'''
import json, sys
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment

rows = json.load(open(sys.argv[1]))
out = sys.argv[2]
wb = Workbook(); ws = wb.active; ws.title = "Outreach"
headers = ["Date","Vendor","Website","Score","Segment","Email","Touch","Sent","Status","Notes"]
ws.append(headers)
hf = PatternFill("solid", fgColor="1F3864")
for c in range(1, len(headers)+1):
    cell = ws.cell(row=1, column=c); cell.font = Font(bold=True, color="FFFFFF"); cell.fill = hf
    cell.alignment = Alignment(horizontal="center")

for nm, web, sc, seg, em, note, live in rows:
    ws.append(["", nm, web, sc, seg, em, "", "", "Not contacted", note])

dv = DataValidation(type="list", allow_blank=True,
    formula1='"Not contacted,Outreach sent,Replied,Consultancy offer sent,Widget installed,Upgraded,Bounced,No contact info,Closed"')
ws.add_data_validation(dv); dv.add("I2:I%d" % (1+len(rows)))
ws.freeze_panes = "A2"
for col,w in {"A":12,"B":24,"C":32,"D":8,"E":30,"F":30,"G":8,"H":8,"I":20,"J":28}.items():
    ws.column_dimensions[col].width = w
wb.save(out)
print("WROTE", out, "rows", len(rows))
'''
open('/tmp/build_tracker_ox.py','w').write(ox_script)
json.dump(rows, open('/tmp/final_rows.json','w'))
out = "/Users/time4you/viralpeps/outreach/trustscore-outreach-tracker.xlsx"
subprocess.run(["bash","-lc",f"cd /Users/time4you/.hermes/viralpeps && source .venv-pdf/bin/activate 2>/dev/null; uv run --with openpyxl python3 /tmp/build_tracker_ox.py /tmp/final_rows.json {out}"], check=True)
emails_filled = sum(1 for r in rows if r[4])
print(f"Emails filled: {emails_filled}/{len(rows)}")
print("Output:", out)

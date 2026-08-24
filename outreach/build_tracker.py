#!/usr/bin/env python3
"""Build TrustScore outreach tracker .xlsx with Status dropdown."""
import json, csv, subprocess, os

# compute scores (same logic as trust-score.ts)
def entity_points(entityType, paymentMethod):
    if paymentMethod in ("card", "bank"): return 25
    if entityType == "ltd": return 20
    if entityType == "sole_trader": return 10
    return 0

def score(vendor):
    auto = vendor.get("_autoChecks") or {}
    s = 0
    if auto.get("coa") is True: s += 25
    if auto.get("ruo") is True: s += 5
    if auto.get("reviews") is True: s += 10
    if auto.get("shipping") is True: s += 5
    if auto.get("contact") is True: s += 10
    if vendor.get("embedded") is True or vendor.get("domainVerified") is True: s += 20
    ep = entity_points(vendor.get("entityType"), vendor.get("paymentMethod"))
    if ep > 0: s += ep
    return min(s, 100)

d = json.load(open("/Users/time4you/viralpeps/src/data/vendors.json"))
vendors = d if isinstance(d, list) else d.get("vendors", d)

data = []
for x in vendors:
    live = (x.get("_autoChecks") or {}).get("live")
    data.append({
        "name": x.get("name", ""), "website": x.get("website", ""),
        "score": score(x), "live": bool(live),
    })
# sort: live/top-score first; offline at bottom
data.sort(key=lambda r: (0 if r["live"] else 1, -r["score"]))

# ---- write via a python one-liner using openpyxl under uv ----
script = r'''
import json, sys
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment

data = json.load(open(sys.argv[1]))
out = sys.argv[2]

wb = Workbook()
ws = wb.active
ws.title = "Outreach"

headers = ["Date","Vendor","Website","Score","Segment","Email","Touch","Sent","Status","Notes"]
ws.append(headers)
hfill = PatternFill("solid", fgColor="1F3864")
for c in range(1, len(headers)+1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = hfill
    cell.alignment = Alignment(horizontal="center")

for r in data:
    ws.append(["", r["name"], r["website"], r["score"], "", "", "", "", "Not contacted", ""])

# Status dropdown
dv = DataValidation(
    type="list",
    formula1='"Not contacted,Outreach sent,Replied,Consultancy offer sent,Widget installed,Upgraded,Bounced,No contact info,Closed"',
    allow_blank=True,
)
ws.add_data_validation(dv)
# apply to rows 2..(1+len)
dv.add("I2:I%d" % (1 + len(data)))

# Column widths
widths = {"A":12,"B":24,"C":34,"D":8,"E":16,"F":26,"G":8,"H":8,"I":20,"J":30}
for col,w in widths.items():
    ws.column_dimensions[col].width = w

# freeze header
ws.freeze_panes = "A2"
wb.save(out)
print("WROTE", out, "rows", len(data))
'''

open("/tmp/build_tracker_ox.py", "w").write(script)
import json
json.dump(data, open("/tmp/tracker_data.json","w"))
os.system("cd /Users/time4you/.hermes/viralpeps && source .venv-pdf/bin/activate 2>/dev/null; "
          "uv run --with openpyxl python3 /tmp/build_tracker_ox.py /tmp/tracker_data.json "
          "/Users/time4you/viralpeps/outreach/trustscore-outreach-tracker.xlsx")
